from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from baseclasses.templatetags.colors import get_color


class MontrekExcelFormatter:
    # Constants for column width calculation
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 50
    COLUMN_PADDING = 2
    BOLD_FONT_MULTIPLIER = 1.2
    NORMAL_FONT_MULTIPLIER = 1.1

    @classmethod
    def format_excel(cls, writer, sheet_name="Sheet1", col_formats=None):
        """Public API — supports both class-level calls (legacy) and instance calls.
        Uses cls() so subclasses inherit correct behavior when called as MyFormatter.format_excel(...).
        """
        cls()._format_excel_impl(writer, sheet_name, col_formats)

    def _format_excel_impl(self, writer, sheet_name="Sheet1", col_formats=None):
        """Format an Excel worksheet with styled headers, alternating rows, and auto-sized columns."""
        worksheet = writer.sheets[sheet_name]
        self._apply_cell_styles(worksheet, col_formats or {})
        self._adjust_column_widths(worksheet)

    def _get_style_objects(self):
        """Create and return all style objects needed for formatting."""
        primary_light = get_color("primary_light").lstrip("#").upper()
        primary = get_color("primary").lstrip("#").upper()

        return {
            "header_fill": PatternFill(
                start_color=primary, end_color=primary, fill_type="solid"
            ),
            "header_font": Font(color="FFFFFF", bold=True),
            "even_row_fill": PatternFill(
                start_color=primary_light, end_color=primary_light, fill_type="solid"
            ),
            "odd_row_fill": PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            ),
            "thin_border": Border(bottom=Side(style="thin", color="E0E0E0")),
        }

    def _apply_cell_styles(self, worksheet, col_formats: dict[int, str | None]):
        """Apply styles to all cells in the worksheet."""
        styles = self._get_style_objects()

        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row):
                if row_idx == 1:
                    self._style_header_cell(cell, styles, col_idx)
                else:
                    self._style_data_cell(cell, row_idx, styles, row)
                    self._format_data_cell(cell, col_formats.get(col_idx))

    def _style_header_cell(self, cell, styles, col_idx: int) -> None:
        """Apply styling to a header cell."""
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="left")
        cell.border = styles["thin_border"]

    def _style_data_cell(self, cell, row_idx, styles, row: list):
        """Apply styling to a data cell with alternating row colors."""
        cell.fill = (
            styles["even_row_fill"] if row_idx % 2 == 0 else styles["odd_row_fill"]
        )
        cell.border = styles["thin_border"]

    def _format_data_cell(self, cell, excel_format_str: str | None):
        """Apply number format and alignment to a data cell."""
        if excel_format_str is not None:
            cell.number_format = excel_format_str
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")

    def _adjust_column_widths(self, worksheet):
        """Auto-size columns based on content with reasonable bounds."""
        for col in worksheet.columns:
            col_cells = list(col)
            if not col_cells:
                continue

            column_letter = get_column_letter(col_cells[0].column)
            max_width = self._calculate_column_width(col_cells)
            worksheet.column_dimensions[column_letter].width = max_width

    def _calculate_column_width(self, col_cells):
        """Calculate the optimal width for a column based on its cells."""
        max_length = 0

        for cell in col_cells:
            if cell.value is not None:
                display_length = self._get_display_length(cell)
                max_length = max(max_length, display_length)

        # Apply bounds and padding
        return min(
            max(
                max_length + self.COLUMN_PADDING,
                self.MIN_COLUMN_WIDTH,
            ),
            self.MAX_COLUMN_WIDTH,
        )

    def _get_display_length(self, cell):
        """Calculate the display length of a cell's content."""
        value = cell.value

        # Get base length based on value type
        if isinstance(value, float):
            base_length = len(f"{value:,.2f}")
        else:
            base_length = len(str(value))

        # Apply font multiplier
        multiplier = (
            self.BOLD_FONT_MULTIPLIER if cell.row == 1 else self.NORMAL_FONT_MULTIPLIER
        )

        return base_length * multiplier
