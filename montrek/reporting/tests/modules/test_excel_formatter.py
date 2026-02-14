from django.test import TestCase
from unittest.mock import Mock
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Font

from reporting.modules.excel_formatter import MontrekExcelFormatter


class MontrekExcelFormatterTests(TestCase):
    """Test suite for MontrekExcelFormatter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

        # Create a mock writer object
        self.mock_writer = Mock()
        self.mock_writer.sheets = {"Sheet1": self.worksheet}

    def tearDown(self):
        """Clean up after tests."""
        self.workbook.close()

    # ==================== Tests for format_excel ====================

    def test_format_excel_applies_styles_and_widths(self):
        """Test that format_excel calls both styling and width adjustment methods."""
        # Add some test data
        self.worksheet["A1"] = "Header"
        self.worksheet["A2"] = "Data"

        MontrekExcelFormatter.format_excel(self.mock_writer, "Sheet1")

        # Verify header cell was styled
        self.assertIsNotNone(self.worksheet["A1"].fill)
        self.assertIsNotNone(self.worksheet["A1"].font)

        # Verify column width was adjusted
        self.assertGreater(self.worksheet.column_dimensions["A"].width, 0)

    # ==================== Tests for _get_style_objects ====================

    def test_get_style_objects_returns_all_styles(self):
        """Test that _get_style_objects returns a dictionary with all required styles."""
        styles = MontrekExcelFormatter._get_style_objects()

        self.assertIn("header_fill", styles)
        self.assertIn("header_font", styles)
        self.assertIn("even_row_fill", styles)
        self.assertIn("odd_row_fill", styles)
        self.assertIn("thin_border", styles)

        self.assertIsInstance(styles["header_fill"], PatternFill)
        self.assertIsInstance(styles["header_font"], Font)
        self.assertIsInstance(styles["even_row_fill"], PatternFill)
        self.assertIsInstance(styles["odd_row_fill"], PatternFill)
        self.assertIsInstance(styles["thin_border"], Border)

    def test_get_style_objects_creates_valid_styles(self):
        """Test that style objects are created with valid properties."""
        styles = MontrekExcelFormatter._get_style_objects()

        # Verify header font is bold and white
        self.assertTrue(styles["header_font"].bold)
        self.assertEqual(styles["header_font"].color.rgb, "00FFFFFF")

        # Verify fills have solid fill type
        self.assertEqual(styles["header_fill"].fill_type, "solid")
        self.assertEqual(styles["even_row_fill"].fill_type, "solid")
        self.assertEqual(styles["odd_row_fill"].fill_type, "solid")

    # ==================== Tests for _apply_cell_styles ====================

    def test_apply_cell_styles_formats_header_row(self):
        """Test that header row (row 1) gets header styling."""
        self.worksheet["A1"] = "Header 1"
        self.worksheet["B1"] = "Header 2"

        MontrekExcelFormatter._apply_cell_styles(self.worksheet)

        # Check header cells have bold white font
        self.assertTrue(self.worksheet["A1"].font.bold)
        self.assertEqual(self.worksheet["A1"].font.color.rgb, "00FFFFFF")
        self.assertTrue(self.worksheet["B1"].font.bold)

    def test_apply_cell_styles_alternates_row_colors(self):
        """Test that data rows alternate between even and odd fill colors."""
        self.worksheet["A1"] = "Header"
        self.worksheet["A2"] = "Row 2 (even)"
        self.worksheet["A3"] = "Row 3 (odd)"
        self.worksheet["A4"] = "Row 4 (even)"

        MontrekExcelFormatter._apply_cell_styles(self.worksheet)

        # Row 2 (even) and Row 4 (even) should have the same fill
        self.assertEqual(
            self.worksheet["A2"].fill.start_color.rgb,
            self.worksheet["A4"].fill.start_color.rgb,
        )

        # Row 3 (odd) should have different fill
        self.assertNotEqual(
            self.worksheet["A2"].fill.start_color.rgb,
            self.worksheet["A3"].fill.start_color.rgb,
        )

    # ==================== Tests for _style_header_cell ====================

    def test_style_header_cell_applies_correct_formatting(self):
        """Test that header cells get the correct fill, font, alignment, and border."""
        cell = self.worksheet["A1"]
        cell.value = "Test Header"
        styles = MontrekExcelFormatter._get_style_objects()

        MontrekExcelFormatter._style_header_cell(cell, styles)

        self.assertEqual(cell.fill, styles["header_fill"])
        self.assertEqual(cell.font, styles["header_font"])
        self.assertEqual(cell.alignment.horizontal, "left")
        self.assertEqual(cell.border, styles["thin_border"])

    # ==================== Tests for _style_data_cell ====================

    def test_style_data_cell_even_row(self):
        """Test that even row data cells get the correct fill."""
        cell = self.worksheet["A2"]
        cell.value = "Test Data"
        styles = MontrekExcelFormatter._get_style_objects()

        MontrekExcelFormatter._style_data_cell(cell, 2, styles)

        self.assertEqual(cell.fill, styles["even_row_fill"])

    def test_style_data_cell_odd_row(self):
        """Test that odd row data cells get the correct fill."""
        cell = self.worksheet["A3"]
        cell.value = "Test Data"
        styles = MontrekExcelFormatter._get_style_objects()

        MontrekExcelFormatter._style_data_cell(cell, 3, styles)

        self.assertEqual(cell.fill, styles["odd_row_fill"])

    def test_style_data_cell_float_formatting(self):
        """Test that float values get proper number formatting and right alignment."""
        cell = self.worksheet["A2"]
        cell.value = 1234.56
        styles = MontrekExcelFormatter._get_style_objects()

        MontrekExcelFormatter._style_data_cell(cell, 2, styles)

        self.assertEqual(cell.number_format, "#,##0.00")
        self.assertEqual(cell.alignment.horizontal, "right")

    def test_style_data_cell_text_alignment(self):
        """Test that non-float values get left alignment."""
        cell = self.worksheet["A2"]
        cell.value = "Text Value"
        styles = MontrekExcelFormatter._get_style_objects()

        MontrekExcelFormatter._style_data_cell(cell, 2, styles)

        self.assertEqual(cell.alignment.horizontal, "left")

    # ==================== Tests for _adjust_column_widths ====================

    def test_adjust_column_widths_sets_width(self):
        """Test that column widths are adjusted based on content."""
        self.worksheet["A1"] = "Short"
        self.worksheet["B1"] = "Very Long Header Text"

        MontrekExcelFormatter._adjust_column_widths(self.worksheet)

        # Column B should be wider than column A
        self.assertGreater(
            self.worksheet.column_dimensions["B"].width,
            self.worksheet.column_dimensions["A"].width,
        )

    def test_adjust_column_widths_handles_empty_columns(self):
        """Test that empty columns are handled gracefully."""
        self.worksheet["A1"] = "Data"
        # Column B is empty

        # Should not raise an exception
        try:
            MontrekExcelFormatter._adjust_column_widths(self.worksheet)
        except Exception as e:
            self.fail(f"_adjust_column_widths raised an exception: {e}")

    # ==================== Tests for _calculate_column_width ====================

    def test_calculate_column_width_respects_min_width(self):
        """Test that calculated width respects minimum width constraint."""
        cell = self.worksheet["A1"]
        cell.value = "A"  # Very short content

        width = MontrekExcelFormatter._calculate_column_width([cell])

        self.assertGreaterEqual(width, MontrekExcelFormatter.MIN_COLUMN_WIDTH)

    def test_calculate_column_width_respects_max_width(self):
        """Test that calculated width respects maximum width constraint."""
        cell = self.worksheet["A1"]
        cell.value = "A" * 100  # Very long content

        width = MontrekExcelFormatter._calculate_column_width([cell])

        self.assertLessEqual(width, MontrekExcelFormatter.MAX_COLUMN_WIDTH)

    def test_calculate_column_width_with_empty_cells(self):
        """Test that empty cells are handled correctly."""
        cell1 = self.worksheet["A1"]
        cell1.value = None
        cell2 = self.worksheet["A2"]
        cell2.value = "Data"

        width = MontrekExcelFormatter._calculate_column_width([cell1, cell2])

        # Should still calculate based on non-empty cell
        self.assertEqual(width, MontrekExcelFormatter.MIN_COLUMN_WIDTH)

    def test_calculate_column_width_includes_padding(self):
        """Test that calculated width includes padding."""
        cell = self.worksheet["A1"]
        cell.value = "Test"

        width = MontrekExcelFormatter._calculate_column_width([cell])

        # Width should be greater than just the content due to padding
        base_length = len("Test") * MontrekExcelFormatter.BOLD_FONT_MULTIPLIER
        self.assertGreater(width, base_length)

    # ==================== Tests for _get_display_length ====================

    def test_get_display_length_for_float(self):
        """Test that float values are formatted with commas and decimals."""
        cell = self.worksheet["A2"]
        cell.value = 1234567.89

        display_length = MontrekExcelFormatter._get_display_length(cell)

        # "1,234,567.89" = 12 characters * 1.1 = 13.2
        expected_length = (
            len("1,234,567.89") * MontrekExcelFormatter.NORMAL_FONT_MULTIPLIER
        )
        self.assertAlmostEqual(display_length, expected_length, places=1)

    def test_get_display_length_for_text(self):
        """Test that text values use their string length."""
        cell = self.worksheet["A2"]
        cell.value = "Test Text"

        display_length = MontrekExcelFormatter._get_display_length(cell)

        expected_length = (
            len("Test Text") * MontrekExcelFormatter.NORMAL_FONT_MULTIPLIER
        )
        self.assertAlmostEqual(display_length, expected_length, places=1)

    def test_get_display_length_for_header_cell(self):
        """Test that header cells (row 1) use bold font multiplier."""
        cell = self.worksheet["A1"]
        cell.value = "Header"

        display_length = MontrekExcelFormatter._get_display_length(cell)

        expected_length = len("Header") * MontrekExcelFormatter.BOLD_FONT_MULTIPLIER
        self.assertAlmostEqual(display_length, expected_length, places=1)

    def test_get_display_length_for_data_cell(self):
        """Test that data cells (row > 1) use normal font multiplier."""
        cell = self.worksheet["A2"]
        cell.value = "Data"

        display_length = MontrekExcelFormatter._get_display_length(cell)

        expected_length = len("Data") * MontrekExcelFormatter.NORMAL_FONT_MULTIPLIER
        self.assertAlmostEqual(display_length, expected_length, places=1)

    def test_get_display_length_for_integer(self):
        """Test that integer values are converted to strings correctly."""
        cell = self.worksheet["A2"]
        cell.value = 12345

        display_length = MontrekExcelFormatter._get_display_length(cell)

        expected_length = len("12345") * MontrekExcelFormatter.NORMAL_FONT_MULTIPLIER
        self.assertAlmostEqual(display_length, expected_length, places=1)

    # ==================== Integration Tests ====================

    def test_full_formatting_with_mixed_data(self):
        """Integration test with mixed data types."""
        # Create a realistic dataset
        self.worksheet["A1"] = "Name"
        self.worksheet["B1"] = "Amount"
        self.worksheet["C1"] = "Percentage"

        self.worksheet["A2"] = "Product A"
        self.worksheet["B2"] = 1234.56
        self.worksheet["C2"] = 0.75

        self.worksheet["A3"] = "Product B"
        self.worksheet["B3"] = 9876543.21
        self.worksheet["C3"] = 0.25

        MontrekExcelFormatter.format_excel(self.mock_writer, "Sheet1")

        # Verify headers are bold
        self.assertTrue(self.worksheet["A1"].font.bold)
        self.assertTrue(self.worksheet["B1"].font.bold)
        self.assertTrue(self.worksheet["C1"].font.bold)

        # Verify float formatting
        self.assertEqual(self.worksheet["B2"].number_format, "#,##0.00")
        self.assertEqual(self.worksheet["C2"].number_format, "#,##0.00")

        # Verify alignment
        self.assertEqual(self.worksheet["A2"].alignment.horizontal, "left")
        self.assertEqual(self.worksheet["B2"].alignment.horizontal, "right")

        # Verify column widths are set
        self.assertGreater(self.worksheet.column_dimensions["A"].width, 0)
        self.assertGreater(self.worksheet.column_dimensions["B"].width, 0)
        self.assertGreater(self.worksheet.column_dimensions["C"].width, 0)

    def test_formatting_with_empty_worksheet(self):
        """Test that formatting handles empty worksheets gracefully."""
        # Don't add any data
        try:
            MontrekExcelFormatter.format_excel(self.mock_writer, "Sheet1")
        except Exception as e:
            self.fail(f"Formatting empty worksheet raised an exception: {e}")

    # ==================== Edge Case Tests ====================

    def test_handles_negative_floats(self):
        """Test that negative float values are formatted correctly."""
        cell = self.worksheet["A2"]
        cell.value = -1234.56

        display_length = MontrekExcelFormatter._get_display_length(cell)

        # "-1,234.56" should be properly calculated
        self.assertGreater(display_length, 0)

    def test_handles_very_small_floats(self):
        """Test that very small float values are formatted correctly."""
        cell = self.worksheet["A2"]
        cell.value = 0.01

        display_length = MontrekExcelFormatter._get_display_length(cell)

        # "0.01" = 4 characters
        expected_length = len("0.01") * MontrekExcelFormatter.NORMAL_FONT_MULTIPLIER
        self.assertAlmostEqual(display_length, expected_length, places=1)

    def test_handles_special_characters(self):
        """Test that special characters in text are handled correctly."""
        cell = self.worksheet["A2"]
        cell.value = "Test™ © Data®"

        try:
            display_length = MontrekExcelFormatter._get_display_length(cell)
            self.assertGreater(display_length, 0)
        except Exception as e:
            self.fail(f"Special characters raised an exception: {e}")
